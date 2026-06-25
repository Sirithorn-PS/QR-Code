import openpyxl

# Read Excel file
file_path = 'stock eneos 30.05.26.xlsx'
wb = openpyxl.load_workbook(file_path)

# Get sheet names
print("Sheet names:", wb.sheetnames)
print()

# Read first sheet
ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print()

# Get headers and first few rows
headers = []
data = []

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 1:
        headers = [h for h in row if h is not None]
        print("Headers:", headers)
    elif i <= 6:  # First 5 data rows
        data.append(row)
        print(f"Row {i}:", row)
    else:
        break

print()
print("Total rows:", ws.max_row)
